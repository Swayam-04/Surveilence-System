import math
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
from threat_type import TARGET_TYPE_PROPERTIES, determine_target_type
from database import SurveillanceDatabase
import uuid
import io
from datetime import datetime
import csv
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

app = Flask(__name__)
db = SurveillanceDatabase()

class Target:
    def __init__(self, position, velocity, acceleration=None, rcs=0, target_type=None, target_id=None, name=None):
        self.position = position
        self.velocity = velocity
        self.acceleration = acceleration
        self.rcs = rcs
        self.target_type = target_type
        self.target_id = target_id or str(uuid.uuid4())[:8]
        self.name = name or f"Target_{self.target_id}"

def euclidean_distance(p1, p2):
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

def is_moving_towards(target_pos, target_vel, key_point):
    dir_to_key = (key_point[0] - target_pos[0], key_point[1] - target_pos[1])
    dot_product = dir_to_key[0] * target_vel[0] + dir_to_key[1] * target_vel[1]
    return dot_product > 0

def calculate_priority(target, key_points, weight_distance=1.0, weight_threat=100.0, weight_rcs=10.0, weight_type=50.0):
    distance_score = euclidean_distance((0, 0), target.position)
    threat_score = 0

    for kp in key_points:
        if is_moving_towards(target.position, target.velocity, kp):
            threat_score += 1 / (euclidean_distance(target.position, kp) + 1e-5)

    rcs_score = weight_rcs * target.rcs

    weight_multiplier = TARGET_TYPE_PROPERTIES.get(target.target_type, TARGET_TYPE_PROPERTIES["unknown"])[6]
    type_score = weight_type * weight_multiplier

    return weight_distance * distance_score - weight_threat * threat_score - rcs_score - type_score

@app.route('/get_target', methods=["POST"])
def index():
    data = request.get_json()
    weight_distance = data.get("weight_distance", 1.0)
    weight_threat = data.get("weight_threat", 100.0)
    weight_rcs = data.get("weight_rcs", 10.0)
    weight_type = data.get("weight_type", 50.0)

    scores = []
    target_types = []
    best_score = float('inf')
    best_target_index = None

    for idx, target_data in enumerate(data.get("targets", [])):
        position = target_data[0]
        velocity = target_data[1]
        acceleration = target_data[2] if len(target_data) > 2 else None
        rcs = target_data[3] if len(target_data) > 3 else 0
        provided_type = target_data[4] if len(target_data) > 4 else None
        target_id = target_data[5].get('id') if len(target_data) > 5 and isinstance(target_data[5], dict) else None

        target_type = determine_target_type(rcs, velocity, acceleration, provided_type)
        target = Target(position, velocity, acceleration, rcs, target_type, target_id)
        score = calculate_priority(target, data.get("key_points", []), weight_distance, weight_threat, weight_rcs, weight_type)
        scores.append(score)
        target_types.append(target_type)

        # Update target in database if it exists
        if target_id:
            try:
                db.update_target_position(target_id, position, velocity, score)
                # Log attack event for predictive analysis
                threat_level = 1 if score < -50 else (2 if score < -100 else 3)
                db.log_attack_event(target_id, position, target_type, threat_level)
            except Exception as e:
                print(f"Error updating target {target_id} in database: {e}")

        if score < best_score:
            best_score = score
            best_target_index = idx

    return jsonify({
        "best_target_index": best_target_index,
        "scores": scores,
        "target_types": target_types
    })

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/targets', methods=['GET'])
def get_targets():
    """Get all targets from database"""
    target_type = request.args.get('type', 'all')
    
    if target_type == 'active':
        targets = db.get_active_targets()
    elif target_type == 'eliminated':
        targets = db.get_eliminated_targets()
    else:
        active = db.get_active_targets()
        eliminated = db.get_eliminated_targets()
        targets = active + eliminated
    
    return jsonify(targets)

@app.route('/api/targets/<target_id>/history', methods=['GET'])
def get_target_history(target_id):
    """Get complete history of a specific target"""
    history = db.get_target_history(target_id)
    return jsonify(history)

@app.route('/api/targets/<target_id>/eliminate', methods=['POST'])
def eliminate_target(target_id):
    """Eliminate a specific target"""
    data = request.get_json() or {}
    elimination_method = data.get('method', 'manual')
    final_score = data.get('final_score')
    
    db.eliminate_target(target_id, elimination_method, final_score)
    return jsonify({"message": f"Target {target_id} eliminated", "status": "success"})

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """Get overall target statistics"""
    stats = db.get_target_statistics()
    return jsonify(stats)

@app.route('/api/targets', methods=['POST'])
def create_target():
    """Create a new target in the database"""
    data = request.get_json()
    
    if not data or 'position' not in data or 'velocity' not in data:
        return jsonify({"error": "Missing required fields"}), 400
    
    try:
        target_data = {
            'target_id': data.get('target_id', str(uuid.uuid4())[:8]),
            'target_type': data.get('target_type', 'unknown'),
            'position': data['position'],
            'velocity': data['velocity'],
            'acceleration': data.get('acceleration', [0, 0]),
            'rcs': data.get('rcs', 0),
            'name': data.get('name')
        }
        
        target_id = db.add_target(target_data)
        return jsonify({
            "message": "Target created successfully",
            "target_id": target_id,
            "status": "success"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/keypoints', methods=['GET'])
def get_keypoints():
    """Get all key points from database"""
    # This would need to be implemented in the database class
    # For now, return empty list
    return jsonify([])

@app.route('/api/cleanup', methods=['POST'])
def cleanup_database():
    """Clean up old event logs"""
    data = request.get_json() or {}
    days_to_keep = data.get('days', 30)
    
    deleted_count = db.cleanup_old_events(days_to_keep)
    return jsonify({
        "message": f"Cleaned up {deleted_count} old events",
        "deleted_count": deleted_count,
        "status": "success"
    })

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Export all database data to Excel file"""
    try:
        # Prefer pandas if available; otherwise fall back to openpyxl directly
        use_pandas = True
        pd = None
        try:
            import pandas as pd  # type: ignore
        except Exception:
            use_pandas = False
            try:
                from openpyxl import Workbook  # type: ignore
            except Exception:
                return jsonify({
                    "error": "Excel export requires either pandas or openpyxl. Install with: pip install openpyxl",
                    "status": "failed"
                }), 501

        # Get all targets from database
        all_targets = db.get_active_targets() + db.get_eliminated_targets()
        
        # Get statistics
        stats = db.get_target_statistics()
        
        # Create Excel file in memory (pandas or openpyxl)
        output = io.BytesIO()
        if use_pandas and pd is not None:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: All Targets
                if all_targets:
                    targets_rows = []
                    for target in all_targets:
                        targets_rows.append({
                            'Target ID': target.get('target_id', 'N/A'),
                            'Name': target.get('name', 'N/A'),
                            'Type': target.get('target_type', 'N/A'),
                            'Position X': target.get('position_x', 0),
                            'Position Y': target.get('position_y', 0),
                            'Velocity X': target.get('velocity_x', 0),
                            'Velocity Y': target.get('velocity_y', 0),
                            'Acceleration X': target.get('acceleration_x', 0),
                            'Acceleration Y': target.get('acceleration_y', 0),
                            'RCS': target.get('rcs', 0),
                            'Created At': target.get('created_at', 'N/A'),
                            'Eliminated At': target.get('eliminated_at', 'Active'),
                            'Elimination Method': target.get('elimination_method', 'N/A'),
                            'Final Score': target.get('final_score', 'N/A')
                        })
                    pd.DataFrame(targets_rows).to_excel(writer, sheet_name='Targets', index=False)

                # Sheet 2: Statistics
                stats_data = {
                    'Metric': [
                        'Total Targets','Active Targets','Eliminated Targets','Average Time to Elimination (seconds)'
                    ],
                    'Value': [
                        stats.get('total_targets', 0),
                        stats.get('active_targets', 0),
                        stats.get('eliminated_targets', 0),
                        stats.get('avg_time_to_elimination_seconds', 'N/A')
                    ]
                }
                for target_type, count in stats.get('targets_by_type', {}).items():
                    stats_data['Metric'].append(f'{target_type.replace("_", " ").title()} Targets')
                    stats_data['Value'].append(count)
                pd.DataFrame(stats_data).to_excel(writer, sheet_name='Statistics', index=False)

                # Sheet 3: System Info
                system_info = {
                    'Information': ['Report Generated','System','Organization','Security Level','Total Records Exported'],
                    'Value': [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'DRDO Radar Surveillance System',
                        'Defence Research & Development Organisation',
                        'CLASSIFIED',
                        len(all_targets)
                    ]
                }
                pd.DataFrame(system_info).to_excel(writer, sheet_name='System Info', index=False)
            output.seek(0)
        else:
            # Build workbook using openpyxl only
            from openpyxl import Workbook  # type: ignore
            wb = Workbook()
            ws_targets = wb.active
            ws_targets.title = 'Targets'
            ws_targets.append(['Target ID','Name','Type','Position X','Position Y','Velocity X','Velocity Y','Acceleration X','Acceleration Y','RCS','Created At','Eliminated At','Elimination Method','Final Score'])
            for t in all_targets:
                ws_targets.append([
                    t.get('target_id','N/A'), t.get('name','N/A'), t.get('target_type','N/A'),
                    t.get('position_x',0), t.get('position_y',0), t.get('velocity_x',0), t.get('velocity_y',0),
                    t.get('acceleration_x',0), t.get('acceleration_y',0), t.get('rcs',0),
                    t.get('created_at','N/A'), t.get('eliminated_at','Active'), t.get('elimination_method','N/A'), t.get('final_score','N/A')
                ])

            ws_stats = wb.createSheet(title='Statistics') if hasattr(wb, 'createSheet') else wb.create_sheet('Statistics')
            ws_stats.append(['Metric','Value'])
            ws_stats.append(['Total Targets', stats.get('total_targets', 0)])
            ws_stats.append(['Active Targets', stats.get('active_targets', 0)])
            ws_stats.append(['Eliminated Targets', stats.get('eliminated_targets', 0)])
            ws_stats.append(['Average Time to Elimination (seconds)', stats.get('avg_time_to_elimination_seconds', 'N/A')])
            for tt, c in stats.get('targets_by_type', {}).items():
                ws_stats.append([f'{tt.replace("_"," ").title()} Targets', c])

            ws_info = wb.create_sheet('System Info')
            ws_info.append(['Information','Value'])
            ws_info.append(['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            ws_info.append(['System','DRDO Radar Surveillance System'])
            ws_info.append(['Organization','Defence Research & Development Organisation'])
            ws_info.append(['Security Level','CLASSIFIED'])
            ws_info.append(['Total Records Exported', len(all_targets)])

            wb.save(output)
            output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DRDO_Radar_Surveillance_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return jsonify({"error": "Failed to generate report"}), 500

@app.route('/api/export/pdf', methods=['GET'])
def export_pdf():
    """Export all database data to PDF file"""
    try:
        # Get all targets from database
        all_targets = db.get_active_targets() + db.get_eliminated_targets()
        
        # Get statistics
        stats = db.get_target_statistics()
        
        # Create PDF file in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.darkblue, alignment=1)
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=14, textColor=colors.darkblue)
        
        # Build PDF content
        story = []
        
        # Title
        story.append(Paragraph("DRDO RADAR SURVEILLANCE REPORT", title_style))
        story.append(Spacer(1, 20))
        
        # System Information
        story.append(Paragraph("System Information", heading_style))
        story.append(Spacer(1, 12))
        
        system_data = [
            ['Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['System:', 'DRDO Radar Surveillance System'],
            ['Organization:', 'Defence Research & Development Organisation'],
            ['Security Level:', 'CLASSIFIED'],
            ['Total Records:', str(len(all_targets))]
        ]
        
        system_table = Table(system_data, colWidths=[2*inch, 3*inch])
        system_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ]))
        
        story.append(system_table)
        story.append(Spacer(1, 20))
        
        # Statistics
        story.append(Paragraph("Statistics Summary", heading_style))
        story.append(Spacer(1, 12))
        
        stats_data = [
            ['Metric', 'Value'],
            ['Total Targets', str(stats.get('total_targets', 0))],
            ['Active Targets', str(stats.get('active_targets', 0))],
            ['Eliminated Targets', str(stats.get('eliminated_targets', 0))],
            ['Avg Time to Elimination (s)', str(stats.get('avg_time_to_elimination_seconds', 'N/A'))]
        ]
        
        # Add target type breakdown
        targets_by_type = stats.get('targets_by_type', {})
        for target_type, count in targets_by_type.items():
            stats_data.append([f'{target_type.replace("_", " ").title()} Targets', str(count)])
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Targets Data
        if all_targets:
            story.append(Paragraph("Target Details", heading_style))
            story.append(Spacer(1, 12))
            
            targets_data = [['Target ID', 'Name', 'Type', 'Position', 'Status', 'Score']]
            for target in all_targets[:20]:  # Limit to first 20 targets for PDF
                position = f"({target.get('position_x', 0):.1f}, {target.get('position_y', 0):.1f})"
                status = 'Active' if not target.get('eliminated_at') else 'Eliminated'
                score = f"{target.get('final_score', 'N/A')}"
                
                targets_data.append([
                    target.get('target_id', 'N/A'),
                    target.get('name', 'N/A'),
                    target.get('target_type', 'N/A'),
                    position,
                    status,
                    score
                ])
            
            targets_table = Table(targets_data, colWidths=[1*inch, 1.2*inch, 1*inch, 1.2*inch, 0.8*inch, 0.8*inch])
            targets_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(targets_table)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DRDO_Radar_Surveillance_Report_{timestamp}.pdf'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        print(f"Error generating PDF report: {e}")
        return jsonify({"error": "Failed to generate PDF report"}), 500

@app.route('/api/export/csv', methods=['GET'])
def export_csv():
    """Export all database data to CSV file"""
    try:
        # Get all targets from database
        all_targets = db.get_active_targets() + db.get_eliminated_targets()
        
        # Create CSV file in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Target ID', 'Name', 'Type', 'Position X', 'Position Y',
            'Velocity X', 'Velocity Y', 'Acceleration X', 'Acceleration Y',
            'RCS', 'Created At', 'Eliminated At', 'Elimination Method', 'Final Score'
        ])
        
        # Write data
        for target in all_targets:
            writer.writerow([
                target.get('target_id', 'N/A'),
                target.get('name', 'N/A'),
                target.get('target_type', 'N/A'),
                target.get('position_x', 0),
                target.get('position_y', 0),
                target.get('velocity_x', 0),
                target.get('velocity_y', 0),
                target.get('acceleration_x', 0),
                target.get('acceleration_y', 0),
                target.get('rcs', 0),
                target.get('created_at', 'N/A'),
                target.get('eliminated_at', 'Active'),
                target.get('elimination_method', 'N/A'),
                target.get('final_score', 'N/A')
            ])
        
        # Convert to bytes
        csv_data = output.getvalue()
        output_bytes = io.BytesIO(csv_data.encode('utf-8'))
        output_bytes.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DRDO_Radar_Surveillance_Report_{timestamp}.csv'
        
        return send_file(
            output_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        print(f"Error generating CSV report: {e}")
        return jsonify({"error": "Failed to generate CSV report"}), 500

@app.route('/api/predictions', methods=['GET'])
def get_predictions():
    """Get predictive analysis of attack patterns"""
    try:
        predictions = db.get_attack_predictions()
        return jsonify(predictions)
    except Exception as e:
        print(f"Error getting predictions: {e}")
        return jsonify({"error": "Failed to get predictions"}), 500

@app.route('/api/sector-heatmap', methods=['GET'])
def get_sector_heatmap():
    """Get sector heatmap data for visualization"""
    try:
        heatmap_data = db.get_sector_heatmap_data()
        return jsonify(heatmap_data)
    except Exception as e:
        print(f"Error getting sector heatmap: {e}")
        return jsonify({"error": "Failed to get sector heatmap"}), 500


# Serve images from local 'image' directory
@app.route('/image/<path:filename>')
def serve_image(filename):
    return send_from_directory('image', filename)

if __name__ == '__main__':
    app.run(debug=True)
